# -*- coding: utf-8 -*-
"""
2024 CUMCM A 题"板凳龙" —— 问题 2：碰撞检测（盘入终止时刻）

模型（与 problem1.md 一致）：
  螺线  r = b·theta,  b = 0.55/(2·pi)
  龙头前把手速度恒 1 m/s，任意时刻 t 走过弧长 s(t)=t
  龙头初始极角 theta0 = 32·pi（第 16 圈）
  逆推法：以已知把手 P_i 为圆心、两孔中心距 d_i 为半径作圆，
          取圆与螺线交点中极角最接近 theta_i 且位于龙头后方者 P_{i+1}

碰撞检测：
  每节板凳视为旋转矩形（板长 L、板宽 0.30 m），两两做分离轴定理(SAT)相交测试。
  相邻板凳共享把手点（允许点接触），板体其余部分重叠即判定碰撞。
  用二分法求首个碰撞时刻（盘入终止时刻），并输出该时刻全部 224 个把手的位置与速度。

输出：
  附件/result2.xlsx（位置 x、y、速度）
  示意图 paper/figures/fig_碰撞临界时刻.png
"""
import numpy as np
import os

# ---------------------------------------------------------------
# 一、参数与求解器
# ---------------------------------------------------------------
PITCH = 0.55
B = PITCH / (2 * np.pi)          # 螺线参数
THETA0 = 32 * np.pi              # 龙头初始极角（第16圈）
V0 = 1.0                         # 龙头前把手速度 (m/s)
WIDTH = 0.30                     # 板宽
HEAD_L, BODY_L = 3.41, 2.20      # 板长
HOLE_OFF = 0.275                 # 孔距板端
D_HEAD = HEAD_L - 2 * HOLE_OFF   # 2.86
D_BODY = BODY_L - 2 * HOLE_OFF   # 1.65


def pos(theta):
    """螺线上极角 theta 处的点."""
    r = B * theta
    return np.array([r * np.cos(theta), r * np.sin(theta)])


def arc_len(theta):
    """弧长从 0 到 theta: b/2 [theta sqrt(1+theta^2) + asinh(theta)]."""
    return 0.5 * B * (theta * np.sqrt(theta * theta + 1) + np.arcsinh(theta))


F0 = arc_len(THETA0)


def theta_head(t):
    """龙头前把手极角: 解 F0 - F(theta) = t（二分法）."""
    lo, hi = 0.0, THETA0
    for _ in range(80):
        mid = 0.5 * (lo + hi)
        if F0 - arc_len(mid) > t:
            lo = mid
        else:
            hi = mid
    return 0.5 * (lo + hi)


def next_theta(th, d):
    """逆推：由 theta 求下一节点（极角更大侧），解 |P(th+u)-P(th)|=d.
    搜索区间 [0, hi] 自适应扩大，保证内圈大极角增量时根不被截断.
    """
    P = pos(th)
    hi = 0.8
    while np.linalg.norm(pos(th + hi) - P) < d and hi < 3.2:
        hi *= 2.0                # 根超出区间则放大
    lo = 0.0
    for _ in range(80):
        mid = 0.5 * (lo + hi)
        if np.linalg.norm(pos(th + mid) - P) > d:
            hi = mid
        else:
            lo = mid
    return th + 0.5 * (lo + hi)


def all_thetas(t):
    """时刻 t 全部 224 把手极角."""
    ths = np.empty(224)
    ths[0] = theta_head(t)
    for i in range(223):
        ths[i + 1] = next_theta(ths[i], D_HEAD if i == 0 else D_BODY)
    return ths


# ---------------------------------------------------------------
# 二、矩形碰撞检测（宽相 AABB 预筛选 + 精相 顶点包含 ∪ SAT）
# ---------------------------------------------------------------
def bench_rect(i, Ps):
    """第 i 节板凳（连接把手 P_i 与 P_{i+1}）的旋转矩形.
    返回 (center, unit_u, half_len, half_wid).
    """
    p0, p1 = Ps[i], Ps[i + 1]
    L = HEAD_L if i == 0 else BODY_L
    v = p1 - p0
    u = v / np.linalg.norm(v)
    c = 0.5 * (p0 + p1)
    return c, u, L / 2.0, WIDTH / 2.0


def rect_corners(r):
    """矩形四个顶点 (N,2): 沿 u/n 方向各 ±half."""
    c, u, hlen, hwid = r
    n = np.array([-u[1], u[0]])
    return np.array([
        c + hlen * u + hwid * n,
        c + hlen * u - hwid * n,
        c - hlen * u + hwid * n,
        c - hlen * u - hwid * n,
    ])


# ---- 宽相（Broad phase）：AABB 外接包围盒，快速剔除远离对 ----
def aabb(r):
    """矩形的外接 AABB: (xmin, ymin, xmax, ymax)."""
    c, u, hlen, hwid = r
    n = np.array([-u[1], u[0]])
    # 半宽在 x、y 方向的分量
    ex = hlen * abs(u[0]) + hwid * abs(n[0])
    ey = hlen * abs(u[1]) + hwid * abs(n[1])
    return c[0] - ex, c[1] - ey, c[0] + ex, c[1] + ey


def aabb_may_overlap(bb1, bb2, gap=1e-6):
    """两个 AABB 是否可能重叠（含 gap 容差）."""
    x1, y1, x2, y2 = bb1
    X1, Y1, X2, Y2 = bb2
    return not (x1 > X2 + gap or X1 > x2 + gap or y1 > Y2 + gap or Y1 > y2 + gap)


# ---- 精相（Narrow phase）：顶点包含 ∪ 分离轴定理(SAT) ----
def _project_interval(c, u, hlen, hwid, axis):
    n = np.array([-u[1], u[0]])
    radius = hlen * abs(np.dot(u, axis)) + hwid * abs(np.dot(n, axis))
    proj = np.dot(c, axis)
    return proj - radius, proj + radius


def _point_in_rect(p, r, tol=1e-9):
    """点 p 是否在旋转矩形 r 内部（含边界）."""
    c, u, hlen, hwid = r
    n = np.array([-u[1], u[0]])
    du = np.dot(p - c, u)
    dn = np.dot(p - c, n)
    return abs(du) <= hlen + tol and abs(dn) <= hwid + tol


def rects_collide(r1, r2, tol=1e-4):
    """两个旋转矩形是否相交（宽相已预筛选；此处为精相）.
    采用两类判据（缺陷1修正：显式避免"一矩形包含另一矩形"漏检）：
      A) 顶点包含：r1 的任一顶点在 r2 内，或 r2 的任一顶点在 r1 内；
      B) 分离轴定理 SAT：4 个轴上投影均重叠（含 A 未覆盖的边-边穿插）。
    任一成立即判定碰撞。SAT 天然覆盖顶点包含，但显式顶点检测作为双重保险。
    """
    # A) 顶点包含（最严重漏检情形：窄/小矩形整体落入另一矩形内部）
    for p in rect_corners(r1):
        if _point_in_rect(p, r2, tol):
            return True
    for p in rect_corners(r2):
        if _point_in_rect(p, r1, tol):
            return True
    # B) 分离轴定理（覆盖其余所有相交情形）
    axes = [r1[1], np.array([-r1[1][1], r1[1][0]]),
            r2[1], np.array([-r2[1][1], r2[1][0]])]
    for ax in axes:
        lo1, hi1 = _project_interval(r1[0], r1[1], r1[2], r1[3], ax)
        lo2, hi2 = _project_interval(r2[0], r2[1], r2[2], r2[3], ax)
        if hi1 - lo2 < tol or hi2 - lo1 < tol:
            return False          # 该轴分离 → 无碰撞
    return True


def detect_collision(ths, first_only=True):
    """检测某时刻是否发生碰撞. 返回 (是否碰撞, 碰撞对列表).
    流程（缺陷2优化）：宽相 AABB 预筛 → 精相（顶点包含 ∪ SAT）。
    相邻板凳（编号差=1）通过把手铰接，允许在把手处接触，不判定为碰撞（缺陷3）；
    非相邻板凳（编号差>=2）板体不得相交。
    """
    Ps = np.array([pos(th) for th in ths])
    n = len(ths) - 1            # 板凳数 223
    rects = [bench_rect(i, Ps) for i in range(n)]
    bbs = [aabb(r) for r in rects]
    hits = []
    for i in range(n):
        for j in range(i + 1, n):
            if j - i == 1:      # 相邻板凳：把手铰接，跳过
                continue
            # 宽相：AABB 相距过远则跳过（避免冗余精相计算）
            if not aabb_may_overlap(bbs[i], bbs[j]):
                continue
            if rects_collide(rects[i], rects[j]):
                hits.append((i + 1, j + 1))   # 板凳编号 1..223
                if first_only:
                    return True, hits
    return bool(hits), hits


# ---------------------------------------------------------------
# 2.5 连续碰撞检测 (CCD)：缺陷4 隧道效应修正
# ---------------------------------------------------------------
def _rect_gap_signed(r1, r2):
    """两矩形带符号间隙：>0 为最小分离量（不相交），<=0 为重叠深度（相交）.
    用 SAT 各轴投影求：
      - 若存在某一轴使投影区间分离（sep>0）→ 不相交，间隙 = 该轴分离量（取所有分离轴中的最小值）；
      - 若所有轴都重叠（无分离轴）→ 相交，间隙 = -最大重叠深度.
    """
    axes = [r1[1], np.array([-r1[1][1], r1[1][0]]),
            r2[1], np.array([-r2[1][1], r2[1][0]])]
    sep_max = -np.inf          # 最大分离量（>0 表示存在分离轴）
    sep_positive = np.inf      # 各分离轴中的最小正分离量
    ovl_max = -np.inf          # 最大重叠深度
    for ax in axes:
        lo1, hi1 = _project_interval(r1[0], r1[1], r1[2], r1[3], ax)
        lo2, hi2 = _project_interval(r2[0], r2[1], r2[2], r2[3], ax)
        sep = max(lo2 - hi1, lo1 - hi2)                 # >0 该轴分离
        ovl = min(hi1, hi2) - max(lo1, lo2)             # >0 该轴重叠
        sep_max = max(sep_max, sep)
        if sep > 0:
            sep_positive = min(sep_positive, sep)
        ovl_max = max(ovl_max, ovl)
    if sep_max > 0:             # 存在分离轴 → 不相交
        return sep_positive
    return -ovl_max             # 全部轴重叠 → 相交（负值为重叠深度）


def _R_bench(bench):
    """板凳外接圆半径（用于 CCD 保守推进的停止上界）."""
    # bench: (c, u, half_len, half_wid)
    return np.hypot(bench[2], bench[3])


def ccd_pair(t0, t1, i, j, iters=28):
    """连续碰撞检测（CCD）：求板凳 i 与 j（编号 1..223）在时间区间 [t0, t1] 内的
    带符号最小间隙及其对应时刻. 用于捕捉"端点无碰撞、区间内短暂重叠"的隧道情形.

    返回 (min_gap, t_contact)：
      min_gap  = 区间内最小带符号间隙（>0 全程分离；<=0 区间内发生/到达碰撞）
      t_contact= 达到最小间隙的时刻.

    用黄金分割（区间内近似单谷）求区间内最小带符号间隙，比固定步长更能定位负间隙.
    """
    def gap_at(t):
        ths = all_thetas(t)
        Ps = np.array([pos(th) for th in ths])
        return _rect_gap_signed(bench_rect(i - 1, Ps), bench_rect(j - 1, Ps))

    @np.vectorize
    def vgap(tv):
        return gap_at(float(tv))

    gr = (np.sqrt(5) - 1) / 2
    def golden_min(a, b, its):
        c = b - gr * (b - a)
        d = a + gr * (b - a)
        fc, fd = gap_at(c), gap_at(d)
        for _ in range(its):
            if fc < fd:
                b, d, fd = d, c, fc
                c = b - gr * (b - a)
                fc = gap_at(c)
            else:
                a, c, fc = c, d, fd
                d = a + gr * (b - a)
                fd = gap_at(d)
        g_min = min(fc, fd)
        return g_min, 0.5 * (a + b)      # (min_gap, t_contact)

    return golden_min(t0, t1, iters)


def ccd_refine_collision(i, j, t_lo, t_hi, tol_t=1e-3, iters=14):
    """对已确认的碰撞对 (i, j) 在区间 [t_lo 无碰撞, t_hi 碰撞] 内，用 CCD 二分逼近
    连续碰撞的精确接触时刻（缺陷4+5：避免隧道漏检并给出精确临界时间）.
    外层用离散 has_coll 快速缩区间（毫秒级精度），对候选区间用 ccd_pair 黄金分割精化间隙.
    返回 t_contact（首次接触时刻）.
    """
    def has_coll(t):
        ths = all_thetas(t)
        c, _ = detect_collision(ths)
        return c
    # 快速离散二分到 tol_t
    while t_hi - t_lo > tol_t:
        t_mid = 0.5 * (t_lo + t_hi)
        if has_coll(t_mid):
            t_hi = t_mid
        else:
            t_lo = t_mid
    # CCD 黄金分割精化：在 [t_lo, t_hi] 内求 (i,j) 最小间隙位置（检验是否有区间内/更早接触）
    gm, tg = ccd_pair(t_lo, t_hi, i, j, iters=iters)
    # 若黄金分割在区间内找到更早的负间隙接触点，则以 tg 为接触时刻
    return tg if (gm <= 0 and tg < t_hi) else t_hi


def ccd_scan(t_start, t_end, n_seg=40, tol_collide=1e-3):
    """对时间段 [t_start, t_end] 做 CCD 扫描，返回全局最小间隙（含隧道检测）.
    对每个子区间 [a,b]，仅对宽相可能重叠的候选非相邻对做 CCD（黄金分割求最小间隙）。
    返回 (min_gap_overall, 对, 时刻)。min_gap_overall > 0 表示全程未发生（含区间内）碰撞。
    """
    best_gap = np.inf
    best_info = None
    n = 223
    for k in range(n_seg):
        a = t_start + (t_end - t_start) * k / n_seg
        b = t_start + (t_end - t_start) * (k + 1) / n_seg
        # 端点位置用于宽相筛选
        ths_a = all_thetas(a)
        ths_b = all_thetas(b)
        Ps_a = np.array([pos(th) for th in ths_a])
        Ps_b = np.array([pos(th) for th in ths_b])
        rects_a = [bench_rect(i, Ps_a) for i in range(n)]
        rects_b = [bench_rect(i, Ps_b) for i in range(n)]
        bbs_a = [aabb(r) for r in rects_a]
        bbs_b = [aabb(r) for r in rects_b]
        for i in range(1, n + 1):
            for j in range(i + 1, n + 1):
                if j - i == 1:
                    continue
                # 宽相：两端点 AABB 均不相交则跳过（该区间内几乎不可能接近）
                if (not aabb_may_overlap(bbs_a[i - 1], bbs_a[j - 1])) and \
                   (not aabb_may_overlap(bbs_b[i - 1], bbs_b[j - 1])):
                    continue
                gm, tg = ccd_pair(a, b, i, j)
                if gm < best_gap:
                    best_gap = gm
                    best_info = (i, j, tg)
    return best_gap, best_info


# ---------------------------------------------------------------
# 三、二分搜索盘入终止时刻
# ---------------------------------------------------------------
def find_stop_time(t_lo=0.0, t_hi=1000.0, tol=1e-3, verbose=True):
    """二分搜索：首次碰撞时刻 = "恰发生碰撞"的临界时刻.
    区间不变式：t_lo 无碰撞、t_hi 有碰撞；收敛后返回 t_hi（首次碰撞上界）。
    缺陷5：用二分法在碰撞区间内逼近精确临界时刻（而非仅"最后无碰撞时刻"）。
    """
    def has_coll(t):
        ths = all_thetas(t)
        c, _ = detect_collision(ths)
        return c

    # 先用大步长找有碰撞的上界
    step = 20.0
    t = 0.0
    while not has_coll(t):
        t += step
        if t > t_hi:
            raise RuntimeError(f"在 {t_hi} s 内未检测到碰撞，请扩大搜索范围")
    t_hi = t
    t_lo = max(0.0, t - step)

    # 二分（t_lo 无碰撞、t_hi 有碰撞）
    while t_hi - t_lo > tol:
        t_mid = 0.5 * (t_lo + t_hi)
        if has_coll(t_mid):
            t_hi = t_mid
        else:
            t_lo = t_mid
    t_stop = t_hi                      # 首次碰撞（恰发生碰撞）时刻
    if verbose:
        print(f"[碰撞检测] 盘入终止时刻 t* = {t_stop:.4f} s")
        ths = all_thetas(t_stop)
        c, hits = detect_collision(ths, first_only=False)
        print(f"[碰撞检测] 碰撞板凳对: {hits}")
    return t_stop


# ---------------------------------------------------------------
# 四、主流程
# ---------------------------------------------------------------
def main():
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    xlsx = os.path.join(root, "附件", "result2.xlsx")
    figpath = os.path.join(root, "paper", "figures", "fig_碰撞临界时刻.png")

    print("=" * 60)
    print("问题 2：碰撞检测（盘入终止时刻）")
    print("=" * 60)
    print("参数: b=%.6f, theta0=%.2f, d_head=%.2f, d_body=%.2f, 板宽=%.2f"
          % (B, THETA0, D_HEAD, D_BODY, WIDTH))

    t_stop = find_stop_time()

    # CCD 连续碰撞精化（缺陷4+5）：对首次碰撞对做区间内连续检测，给出更精确接触时刻
    try:
        c_, hits_ = detect_collision(all_thetas(t_stop), first_only=False)
        if hits_:
            ip, jp = hits_[0]
            t_ccd = ccd_refine_collision(ip, jp, t_stop - 0.5, t_stop + 0.5, tol_t=1e-3, iters=12)
            print(f"[CCD] 连续碰撞精化: 对 ({ip},{jp}) 接触时刻 ≈ {t_ccd:.5f} s (二分 t*={t_stop:.4f})")
            t_stop = min(t_stop, t_ccd)   # 采用 CCD 更精确（更早）的接触时刻
    except Exception as e:
        print(f"[CCD] 精化跳过: {e}")


    # 终止时刻的全部数据
    ths = all_thetas(t_stop)
    Ps = np.array([pos(th) for th in ths])
    # 速度：速度分解法（刚体铰接链递推，与 problem1 一致，更精确）
    def _tan(th):
        dr = np.array([np.cos(th) - th * np.sin(th), np.sin(th) + th * np.cos(th)]) * B
        return dr / np.linalg.norm(dr)
    vels = np.empty(224)
    vels[0] = 1.0
    for i in range(223):
        e = Ps[i + 1] - Ps[i]
        e = e / np.linalg.norm(e)
        vels[i + 1] = vels[i] * (np.dot(_tan(ths[i]), e) / np.dot(_tan(ths[i + 1]), e))

    print(f"\n终止时刻 t* = {t_stop:.4f} s")
    print("龙头: 位置({:.6f}, {:.6f}), 速度 {:.6f}".format(*Ps[0], vels[0]))
    print("龙尾后把手: 位置({:.6f}, {:.6f}), 速度 {:.6f}".format(*Ps[223], vels[223]))

    # 写入 result2.xlsx
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx)
        ws = wb["Sheet1"]
        for i in range(224):
            ws.cell(row=2 + i, column=2, value=round(float(Ps[i][0]), 6))
            ws.cell(row=2 + i, column=3, value=round(float(Ps[i][1]), 6))
            ws.cell(row=2 + i, column=4, value=round(float(vels[i]), 6))
        wb.save(xlsx)
        print(f"[输出] 已写入 {xlsx}")
    except Exception as e:
        print(f"[警告] 写入 Excel 失败: {e}")

    # 绘制碰撞临界时刻示意图
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "SimSun"]
        plt.rcParams["axes.unicode_minus"] = False
        fig, ax = plt.subplots(figsize=(8, 8))
        ax.plot(Ps[:, 0], Ps[:, 1], color="#2f5f9e", lw=1.2)
        ax.scatter(*Ps[0], color="red", s=50, zorder=5, label="龙头前把手")
        ax.scatter(*Ps[223], color="green", s=40, zorder=5, label="龙尾后把手")
        c, hits = detect_collision(ths, first_only=False)
        for a, b in hits[:8]:
            ax.plot([Ps[a-1][0], Ps[a][0]], [Ps[a-1][1], Ps[a][1]], color="#e07b39", lw=3)
            ax.plot([Ps[b-1][0], Ps[b][0]], [Ps[b-1][1], Ps[b][1]], color="#e07b39", lw=3)
        if hits:
            ax.set_title(f"碰撞临界时刻 t*={t_stop:.3f} s（橙色=发生碰撞的板凳）")
        else:
            ax.set_title(f"t*={t_stop:.3f} s 盘入整体形态")
        ax.set_aspect("equal")
        ax.set_xlabel("x (m)"); ax.set_ylabel("y (m)")
        ax.grid(alpha=0.3)
        ax.legend(fontsize=10)
        plt.tight_layout()
        plt.savefig(figpath, dpi=200)
        plt.close()
        print(f"[输出] 已绘制 {figpath}")
    except Exception as e:
        print(f"[警告] 绘图失败: {e}")


if __name__ == "__main__":
    main()