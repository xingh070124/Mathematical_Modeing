# -*- coding: utf-8 -*-
"""
2024 CUMCM A 题"板凳龙" —— 问题一：求解（细步长）+ 实体碰撞检测

模型（与 problem1.md 一致）：
  螺线 r = b·theta, b = 0.55/(2·pi)
  龙头前把手恒速 1 m/s，t 时刻走过弧长 s(t)=t
  龙头初始极角 theta0 = 32·pi（第16圈）
  逆推法求全部 224 把手位置

改进：
  1) 细步长 dt=0.1s 计算，速度用细步长中心差分（比 1s 差分精度高约 50 倍）；
  2) 实体碰撞检测：每节板凳按矩形实体（板长 L × 板宽 0.30 m）建模，
     用分离轴定理(SAT)检测所有非相邻板凳对是否相交，输出最小间隙随时间变化。

输出：
  附件/result1.xlsx（位置表 0~300s、速度表 0~300s，6 位小数）
  paper/figures/fig_问题一碰撞间隙.png（最小间隙 vs 时间）
  控制台打印关键时刻表
"""
import numpy as np
import os
import time
from scipy.optimize import brentq

# ---------------- 参数 ----------------
PITCH = 0.55
B = PITCH / (2 * np.pi)
THETA0 = 32 * np.pi
V0 = 1.0
HEAD_L, BODY_L = 3.41, 2.20
HOLE_OFF = 0.275
D_HEAD = HEAD_L - 2 * HOLE_OFF   # 2.86
D_BODY = BODY_L - 2 * HOLE_OFF   # 1.65
WIDTH = 0.30                     # 板宽（实体碰撞用）
DT = 0.1                         # 细步长 (s)


def pos(theta):
    r = B * theta
    return np.array([r * np.cos(theta), r * np.sin(theta)])


def arc_len(theta):
    return 0.5 * B * (theta * np.sqrt(theta * theta + 1) + np.arcsinh(theta))


F0 = arc_len(THETA0)


def theta_head(t):
    lo, hi = 0.0, THETA0
    for _ in range(60):
        mid = 0.5 * (lo + hi)
        if F0 - arc_len(mid) > t:
            lo = mid
        else:
            hi = mid
    return 0.5 * (lo + hi)


def _dist_u(u, th, d):
    """|P(th+u) - P(th)| - d，u∈[0,hi]."""
    P = pos(th)
    return np.linalg.norm(pos(th + u) - P) - d


def next_theta(th, d):
    """逆推：由 theta 求下一节点（极角更大侧），brentq 求根.
    因 d>0 且 u=0 时 f=-d<0，只需找 f>0 的 hi 即得括根区间 [0,hi].
    """
    hi = 0.8
    while _dist_u(hi, th, d) < 0 and hi < 3.2:
        hi *= 2.0
    u_root = brentq(lambda u: _dist_u(u, th, d), 0.0, hi, xtol=1e-12)
    return th + u_root


def all_thetas(t):
    ths = np.empty(224)
    ths[0] = theta_head(t)
    for i in range(223):
        ths[i + 1] = next_theta(ths[i], D_HEAD if i == 0 else D_BODY)
    return ths


def all_positions(t):
    return np.array([pos(th) for th in all_thetas(t)])


# ---------------- 实体碰撞检测（SAT，按板凳宽 0.30m） ----------------
def bench_rect(i, Ps):
    """第 i 节板凳（0 基，连接把手 i 与 i+1）的旋转矩形. 返回 (center, unit_u, half_len, half_wid)."""
    p0, p1 = Ps[i], Ps[i + 1]
    L = HEAD_L if i == 0 else BODY_L
    v = p1 - p0
    u = v / np.linalg.norm(v)
    c = 0.5 * (p0 + p1)
    return c, u, L / 2.0, WIDTH / 2.0


def _proj_interval(c, u, hlen, hwid, axis):
    n = np.array([-u[1], u[0]])
    radius = hlen * abs(np.dot(u, axis)) + hwid * abs(np.dot(n, axis))
    p = np.dot(c, axis)
    return p - radius, p + radius


def rects_overlap(r1, r2, tol=1e-6):
    axes = [r1[1], np.array([-r1[1][1], r1[1][0]]),
            r2[1], np.array([-r2[1][1], r2[1][0]])]
    for ax in axes:
        lo1, hi1 = _proj_interval(r1[0], r1[1], r1[2], r1[3], ax)
        lo2, hi2 = _proj_interval(r2[0], r2[1], r2[2], r2[3], ax)
        if hi1 - lo2 < tol or hi2 - lo1 < tol:
            return False
    return True


def min_gap(ths):
    """全部非相邻板凳对的最小间隙（重叠为负，间隙为正，单位 m）. 向量化实现.
    返回 (min_gap, pair)。相邻板凳(编号差=1)经把手铰接，跳过。

    每对矩形：沿 4 个分离轴分别求投影，若某轴投影分离则该对不相交，
    间隙取该对的"最小正分离量"；若 4 轴均重叠则间隙为负（重叠深度）。
    全局 min_gap > 0 表示全程无碰撞。
    """
    Ps = np.array([pos(th) for th in ths])

    p0, p1 = Ps[:-1], Ps[1:]
    L = np.where(np.arange(223) == 0, HEAD_L, BODY_L)
    v = p1 - p0
    nrm = np.linalg.norm(v, axis=1, keepdims=True)
    u = v / nrm
    c = 0.5 * (p0 + p1)
    hlen = L / 2.0
    hwid = np.full(223, WIDTH / 2.0)

    idx = np.triu_indices(223, 1)
    i, j = idx
    mask = (j - i) > 1
    i, j = i[mask], j[mask]
    ci, ui = c[i], u[i]
    cj, uj = c[j], u[j]
    hli, hwi = hlen[i], hwid[i]
    hlj, hwj = hlen[j], hwid[j]
    ni = np.stack([-ui[:, 1], ui[:, 0]], axis=1)
    nj = np.stack([-uj[:, 1], uj[:, 0]], axis=1)

    n_pairs = len(i)
    # 每个轴: sep>0 表示该轴分离(不相交), ovl<0 表示该轴重叠
    sep_all = np.zeros((4, n_pairs))     # 各轴分离量(<=0 表示该轴重叠)
    ovl_all = np.zeros((4, n_pairs))     # 各轴重叠量(>0 表示该轴重叠)
    for k, ax in enumerate([ui, ni, uj, nj]):
        rad_i = hli * np.abs((ui * ax).sum(1)) + hwi * np.abs((ni * ax).sum(1))
        rad_j = hlj * np.abs((uj * ax).sum(1)) + hwj * np.abs((nj * ax).sum(1))
        pci = (ci * ax).sum(1)
        pcj = (cj * ax).sum(1)
        lo1, hi1 = pci - rad_i, pci + rad_i
        lo2, hi2 = pcj - rad_j, pcj + rad_j
        sep_all[k] = np.maximum(lo2 - hi1, lo1 - hi2)      # >0: 该轴分离
        ovl_all[k] = np.minimum(hi1, hi2) - np.maximum(lo1, lo2)  # >0: 该轴重叠

    # 存在任一轴分离 => 不相交, 间隙 = 最小正分离量(取各轴sep>0的min, 否则为0)
    sep_min = np.where(sep_all > 0, sep_all, np.inf).min(axis=0)
    collided = np.isinf(sep_min)                           # 4轴都无分离 => 重叠
    # 相交时间隙 = -最大重叠深度
    ovl_max = ovl_all.max(axis=0)                          # 重叠深度(>0)
    pair_gap = np.where(collided, -ovl_max, sep_min)

    k = int(np.argmin(pair_gap))
    return float(pair_gap[k]), (int(i[k]) + 1, int(j[k]) + 1)


# ---------------- 主流程 ----------------
def main():
    t0 = time.time()
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    xlsx = os.path.join(root, "附件", "result1.xlsx")
    figpath = os.path.join(root, "paper", "figures", "fig_问题一碰撞间隙.png")

    N_s = 301                    # 每秒采样 0~300
    n_small = int(300 / DT) + 1  # 细步长采样点数
    t_small = np.arange(n_small) * DT

    print("=" * 60)
    print("问题一：细步长求解 + 实体碰撞检测")
    print(f"细步长 dt = {DT} s, 细步采样点 {n_small} 个, 输出每秒 {N_s} 个点")
    print("=" * 60)

    # 1) 细步长位置（全部采样点）
    print("[1/3] 细步长位置计算 ...")
    pos_small = np.array([all_positions(t) for t in t_small])   # (n_small, 224, 2)

    # 2) 细步长中心差分速度（端点单侧差分）
    print("[2/3] 细步长速度（中心差分）...")
    vel_small = np.empty_like(pos_small[:, :, 0])
    vel_small[0] = np.linalg.norm(pos_small[1] - pos_small[0], axis=1) / DT
    vel_small[-1] = np.linalg.norm(pos_small[-1] - pos_small[-2], axis=1) / DT
    vel_small[1:-1] = np.linalg.norm(pos_small[2:] - pos_small[:-2], axis=2) / (2 * DT)

    # 3) 实体碰撞检测：全程每秒采样点
    print("[3/3] 实体碰撞检测（板凳宽 0.30 m）...")
    gaps = np.empty(N_s)
    worst_pair = None
    min_gap_all = 1e9
    for k in range(N_s):
        ths = all_thetas(k)
        g, pair = min_gap(ths)
        gaps[k] = g
        if g < min_gap_all:
            min_gap_all = g
            worst_pair = pair
    print(f"   0~300 s 内最小间隙 = {min_gap_all:.4f} m (板凳对 {worst_pair})")
    print(f"   结论: {'全程无碰撞' if min_gap_all > 0 else '发生碰撞!'}")

    # 写入 result1.xlsx（位置 + 速度，6 位小数）
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    ws_pos = wb["位置"]
    ws_vel = wb["速度"]
    # 每秒数据 = 细步长每隔 10 点
    sec_idx = np.arange(0, n_small, int(1 / DT))
    for i in range(224):
        for c in range(301):
            ws_pos.cell(row=2 + 2 * i, column=2 + c, value=round(float(pos_small[sec_idx[c], i, 0]), 6))
            ws_pos.cell(row=3 + 2 * i, column=2 + c, value=round(float(pos_small[sec_idx[c], i, 1]), 6))
            ws_vel.cell(row=2 + i, column=2 + c, value=round(float(vel_small[sec_idx[c], i]), 6))
    wb.save(xlsx)
    print(f"[输出] result1.xlsx 已更新（细步长速度）")

    # 绘图：最小间隙 vs 时间
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm
        for fp in [r"C:\Windows\Fonts\msyh.ttc", r"C:\Windows\Fonts\simhei.ttf"]:
            if os.path.exists(fp):
                fm.fontManager.addfont(fp)
        plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "SimSun"]
        plt.rcParams["axes.unicode_minus"] = False
        fig, ax = plt.subplots(figsize=(9, 4.6))
        ax.plot(np.arange(N_s), gaps, color="#1f6fb2", lw=1.8)
        ax.axhline(0, color="red", ls="--", lw=1.2, label="碰撞阈值 (0)")
        ax.set_xlabel("t (s)")
        ax.set_ylabel("最小板凳间隙 (m)")
        ax.set_title(f"问题一 0~300 s 实体碰撞检测：最小间隙（板宽 0.30 m，最小间隙 {min_gap_all:.3f} m）")
        ax.legend()
        ax.grid(alpha=0.3)
        plt.tight_layout()
        plt.savefig(figpath, dpi=200)
        plt.close()
        print(f"[输出] {figpath}")
    except Exception as e:
        print(f"[警告] 绘图失败: {e}")

    print(f"\n耗时 {time.time() - t0:.1f} s")

    # 关键时刻表（论文用）
    print("\n=== 关键时刻位置 (m) ===")
    keys = [0, 1, 51, 101, 151, 201, 223]
    names = ["龙头", "第1节", "第51节", "第101节", "第151节", "第201节", "龙尾后"]
    for h, nm in zip(keys, names):
        row = " ".join(f"{pos_small[sec_idx[c], h, 0]:.6f},{pos_small[sec_idx[c], h, 1]:.6f}"
                       for c in [0, 60, 120, 180, 240, 300])
        print(f"{nm}: {row}")
    print("\n=== 关键时刻速度 (m/s) ===")
    for h, nm in zip(keys, names):
        row = " ".join(f"{vel_small[sec_idx[c], h]:.6f}" for c in [0, 60, 120, 180, 240, 300])
        print(f"{nm}: {row}")
    print(f"\n最小间隙对: {worst_pair}, 最小间隙: {min_gap_all:.4f} m")


if __name__ == "__main__":
    main()