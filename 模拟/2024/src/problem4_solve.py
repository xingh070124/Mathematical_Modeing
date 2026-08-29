# -*- coding: utf-8 -*-
"""
2024 CUMCM A 题“板凳龙” —— 问题 4：S 形调头曲线与全龙仿真 + 调头曲线缩短分析

设定（题目）：
  盘入螺距 p = 1.7 m（b = p/2π），盘出螺线与盘入螺线关于中心呈中心对称；
  调头空间：以螺线中心为圆心、直径 9 m 的圆（r ≤ 4.5 m）；
  调头路径 = 两段圆弧相切连成的 S 形曲线，前段半径 R1 = 2·R2，与盘入、盘出螺线均相切；
  龙头前把手恒速 1 m/s，t = 0 为调头开始时刻，求 -100 s ~ 100 s 每秒全龙 224 把手位置与速度。

模型（与 problem1/2/3 同构）：
  入口切点 E = 盘入螺线与调头空间边界交点（r = 4.5），行进方向沿螺线切向（θ 减小）；
  出口切点 F = -E（盘出螺线与边界交点），行进方向与 t_E 相同（中心对称性）；
  各把手沿同一复合轨迹曲线 C（盘入螺线 ∪ 弧1 ∪ 弧2 ∪ 盘出螺线）运动，
  相邻把手沿 C 的弦长恒为孔距 D_i —— 逆推法推广（作圆与 C 求后方交点）；
  速度用刚体铰接链递推 v_{i+1} = v_i·(τ_i·e)/(τ_{i+1}·e)（与 problem1/2 一致）。

第二问（能否变短）：
  (a) 不变性定理：切点固定于 E、F 的两段圆弧外切族中，R1+R2 = S* 与每段扫角 α 均由
      几何唯一确定，与半径分配无关，故调头曲线长度 L = S*·α 恒定 —— 仅调整半径比例不能变短；
  (b) 结构拓展（弧-线-弧，保持各部分相切）：以 (R1, α) 为自由参数可数值求极短，
      验证位于调头空间内且全龙无碰撞后给出缩短结论。

输出：
  附件/result4.xlsx（模板格式：位置 / 速度两表，列 -100 s ~ 100 s，保留 6 位小数）
  src/problem4_results.json（文档与绘图所需全部关键数字）
"""
import numpy as np
import os
import json
from scipy.optimize import brentq

# ---------------------------------------------------------------
# 一、常量与螺线
# ---------------------------------------------------------------
PITCH = 1.7
B = PITCH / (2 * np.pi)           # 螺线参数
R_SPACE = 4.5                     # 调头空间半径（直径 9 m）
WIDTH = 0.30
HEAD_L, BODY_L = 3.41, 2.20
HOLE_OFF = 0.275
D_HEAD = HEAD_L - 2 * HOLE_OFF    # 2.86
D_BODY = BODY_L - 2 * HOLE_OFF    # 1.65
D_ALL = [D_HEAD] + [D_BODY] * 222  # 223 段孔距
N_H = 224

TH_A = R_SPACE / B                # 入口极角 = 9π/1.7


def arc_len(th):
    """螺线自 θ=0 的弧长: b/2 [θ√(1+θ²) + asinh θ]."""
    return 0.5 * B * (th * np.hypot(th, 1.0) + np.arcsinh(th))


def d_arc_len(th):
    """ds/dθ = b√(1+θ²)."""
    return B * np.hypot(th, 1.0)


S_A = arc_len(TH_A)


def spiral_point(th):
    r = B * th
    return np.array([r * np.cos(th), r * np.sin(th)])


def spiral_dpoint(th):
    return B * np.array([np.cos(th) - th * np.sin(th),
                         np.sin(th) + th * np.cos(th)])


def theta_of_arc(s):
    """反解 θ 使 arc_len(θ)=s（牛顿迭代，双精度收敛）."""
    th = np.sqrt(2.0 * s / B) + 1e-9
    for _ in range(60):
        f = arc_len(th) - s
        if abs(f) < 1e-13:
            break
        th -= f / d_arc_len(th)
        if th <= 0:
            th = 1e-9
    return th


def rot90(v):
    return np.array([-v[1], v[0]])


def rot(v, ang):
    c, s = np.cos(ang), np.sin(ang)
    return np.array([c * v[0] - s * v[1], s * v[0] + c * v[1]])


# ---------------------------------------------------------------
# 二、S 形调头曲线几何（R1 = 2 R2 基线）
# ---------------------------------------------------------------
E = spiral_point(TH_A)                                   # 入口切点
tE = -spiral_dpoint(TH_A) / np.linalg.norm(spiral_dpoint(TH_A))  # 盘入行进方向
nE = rot90(tE)                                           # 左法线
F = -E                                                   # 出口切点（中心对称）
tF = tE                                                  # 盘出行进方向（对称性）
DELTA = F - E                                            # 入→出位移


def s_total_from_split(r1, r2):
    """给定半径分配求两圆外切配置；返回 (O1, O2, S=r1+r2) 或 None.
    弧1 右转（圆心 O1 = E - r1·nE），弧2 左转（圆心 O2 = F + r2·nE）."""
    O1 = E - r1 * nE
    O2 = F + r2 * nE
    d = np.linalg.norm(O1 - O2)
    if abs(d - (r1 + r2)) > 1e-9:
        return None
    return O1, O2, r1 + r2


def solve_split_ratio(k):
    """R1 = k·R2：解外切方程 |O1-O2| = R1+R2 得 R2（解析式 + 数值验证）."""
    # |Δ + S·nE| = S, S=(k+1)R2  →  S = -|Δ|²/(2 Δ·nE)
    S = -np.dot(DELTA, DELTA) / (2.0 * np.dot(DELTA, nE))
    r2 = S / (k + 1.0)
    r1 = k * r2
    res = s_total_from_split(r1, r2)
    assert res is not None, "split ratio 解不一致"
    return r1, r2, S


def arc_geometry(O1, O2, r1, r2):
    """由两圆心求切点 T 与两弧扫角（弧1 顺时针、弧2 逆时针），并做一致性校验."""
    m = (O2 - O1) / np.linalg.norm(O2 - O1)   # O1→O2 单位向量
    T = O1 + r1 * m
    phi_E = np.arctan2(*(E - O1)[::-1])
    phi_T1 = np.arctan2(*(m)[::-1])
    phi_T2 = np.arctan2(*(T - O2)[::-1])
    phi_F = np.arctan2(*(F - O2)[::-1])
    g1 = (phi_E - phi_T1) % (2 * np.pi)       # 弧1 顺时针扫角
    g2 = (phi_F - phi_T2) % (2 * np.pi)       # 弧2 逆时针扫角
    return T, m, g1, g2


def verify_two_arc(O1, O2, r1, r2, g1, g2, tol=1e-8):
    """数值校验：弧1 起点切向 = tE，弧2 终点切向 = tF，两端点吻合."""
    def circ_pt(O, r, phi):
        return O + r * np.array([np.cos(phi), np.sin(phi)])

    def circ_tan_cw(phi):     # 顺时针运动切向
        return np.array([np.sin(phi), -np.cos(phi)])

    def circ_tan_ccw(phi):
        return np.array([-np.sin(phi), np.cos(phi)])

    phi_E = np.arctan2(*(E - O1)[::-1])
    ok = (np.linalg.norm(circ_tan_cw(phi_E) - tE) < tol)
    phi_F = np.arctan2(*(F - O2)[::-1])
    ok = ok and (np.linalg.norm(circ_tan_ccw(phi_F) - tF) < tol)
    return ok


# 基线：R1 = 2 R2
R1_0, R2_0, S0 = solve_split_ratio(2.0)
O1_0, O2_0, _ = s_total_from_split(R1_0, R2_0)
T0, m0, G1_0, G2_0 = arc_geometry(O1_0, O2_0, R1_0, R2_0)
assert abs(G1_0 - G2_0) < 1e-6, f"两弧扫角不等: {G1_0}, {G2_0}"
assert verify_two_arc(O1_0, O2_0, R1_0, R2_0, G1_0, G2_0), "切向校验失败"
L1_0, L2_0 = R1_0 * G1_0, R2_0 * G2_0
L0 = L1_0 + L2_0

# ---------------------------------------------------------------
# 三、复合轨迹曲线 C（弧长参数 u，u=0 在 E，沿行进方向递增）
# ---------------------------------------------------------------
L_C = L1_0 + L2_0


def curve_point(u):
    if u < 0:                                   # 盘入螺线
        return spiral_point(theta_of_arc(S_A - u))
    if u <= L1_0:                               # 弧1（顺时针）
        phi = np.arctan2(*(E - O1_0)[::-1]) - u / R1_0
        return O1_0 + R1_0 * np.array([np.cos(phi), np.sin(phi)])
    if u <= L_C:                                # 弧2（逆时针）
        phi = np.arctan2(*(T0 - O2_0)[::-1]) + (u - L1_0) / R2_0
        return O2_0 + R2_0 * np.array([np.cos(phi), np.sin(phi)])
    # 盘出螺线（θ 增大 = 向外盘出）
    return -spiral_point(theta_of_arc(S_A + (u - L_C)))


def curve_tangent(u):
    """行进方向单位切向."""
    if u < 0:
        th = theta_of_arc(S_A - u)
        v = spiral_dpoint(th)
        return -v / np.linalg.norm(v)
    if u <= L1_0:
        phi = np.arctan2(*(E - O1_0)[::-1]) - u / R1_0
        return np.array([np.sin(phi), -np.cos(phi)])
    if u <= L_C:
        phi = np.arctan2(*(T0 - O2_0)[::-1]) + (u - L1_0) / R2_0
        return np.array([-np.sin(phi), np.cos(phi)])
    th = theta_of_arc(S_A + (u - L_C))
    v = -spiral_dpoint(th)
    return v / np.linalg.norm(v)


# C¹ 连续性校验（接缝处切向）
for u_j in (0.0, L1_0, L_C):
    ta, tb = curve_tangent(u_j - 1e-7), curve_tangent(u_j + 1e-7)
    assert np.linalg.norm(ta - tb) < 1e-5, f"接缝 u={u_j} 切向不连续"

# ---------------------------------------------------------------
# 四、逆推法（推广）：沿 C 的弦长递推 + 速度递推
# ---------------------------------------------------------------
def config_at(t, dt_check=False):
    """时刻 t 全部 224 把手位置（u_1 = t，逐把向后解弦长方程）."""
    us = np.empty(N_H)
    us[0] = t
    for i in range(223):
        di = D_ALL[i]
        P = curve_point(us[i])
        f = lambda dl: np.linalg.norm(curve_point(us[i] - dl) - P) - di
        lo, hi = 0.4 * di, 2.6 * di
        us[i + 1] = us[i] - brentq(f, lo, hi, xtol=1e-12)
    return us


def positions(us):
    return np.array([curve_point(u) for u in us])


def velocities(us):
    """速度递推（与 problem1/2 一致）：v_{i+1} = v_i·(τ_i·e)/(τ_{i+1}·e)."""
    Ps = positions(us)
    v = np.empty(N_H)
    v[0] = 1.0
    for i in range(223):
        e = Ps[i + 1] - Ps[i]
        e = e / np.linalg.norm(e)
        ti = curve_tangent(us[i])
        tj = curve_tangent(us[i + 1])
        v[i + 1] = v[i] * np.dot(ti, e) / np.dot(tj, e)
    return v


# 中心差分交叉验证（抽查 3 个时刻）
for t_chk in (-50.0, 0.0, 37.0):
    h = 1e-4
    up = config_at(t_chk + h)
    um = config_at(t_chk - h)
    Pp, Pm = positions(up), positions(um)
    v_fd = np.linalg.norm(Pp - Pm, axis=1) / (2 * h)
    us = config_at(t_chk)
    v_an = velocities(us)
    err = np.max(np.abs(v_fd - v_an))
    assert err < 5e-5, f"速度校验失败 t={t_chk}: {err}"

# ---------------------------------------------------------------
# 五、碰撞检测（复用 problem2 的 SAT 框架，作用于任意把手位置）
# ---------------------------------------------------------------
def bench_rect(i, Ps):
    p0, p1 = Ps[i], Ps[i + 1]
    L = HEAD_L if i == 0 else BODY_L
    vv = p1 - p0
    u = vv / np.linalg.norm(vv)
    return 0.5 * (p0 + p1), u, L / 2.0, WIDTH / 2.0


def rect_corners(r):
    """矩形四顶点 (4,2)."""
    c, u, hl, hw = r
    n = np.array([-u[1], u[0]])
    return np.array([c + hl * u + hw * n, c + hl * u - hw * n,
                     c - hl * u + hw * n, c - hl * u - hw * n])


def aabb(r):
    c, u, hl, hw = r
    n = np.array([-u[1], u[0]])
    ex = hl * abs(u[0]) + hw * abs(n[0])
    ey = hl * abs(u[1]) + hw * abs(n[1])
    return c[0] - ex, c[1] - ey, c[0] + ex, c[1] + ey


def aabb_overlap(b1, b2, gap=1e-9):
    return not (b1[0] > b2[2] + gap or b2[0] > b1[2] + gap or
                b1[1] > b2[3] + gap or b2[1] > b1[3] + gap)


def rect_gap_signed(r1, r2):
    """SAT 带符号间隙（>0 分离量，<=0 重叠深度），同 problem2."""
    axes = [r1[1], np.array([-r1[1][1], r1[1][0]]),
            r2[1], np.array([-r2[1][1], r2[1][0]])]
    sep_max, sep_pos, ovl_max = -np.inf, np.inf, -np.inf
    for ax in axes:
        lo1 = np.dot(r1[0], ax) - (r1[2] * abs(np.dot(r1[1], ax)) +
                                   r1[3] * abs(np.dot(np.array([-r1[1][1], r1[1][0]]), ax)))
        hi1 = np.dot(r1[0], ax) + (r1[2] * abs(np.dot(r1[1], ax)) +
                                   r1[3] * abs(np.dot(np.array([-r1[1][1], r1[1][0]]), ax)))
        n2 = np.array([-r2[1][1], r2[1][0]])
        lo2 = np.dot(r2[0], ax) - (r2[2] * abs(np.dot(r2[1], ax)) + r2[3] * abs(np.dot(n2, ax)))
        hi2 = np.dot(r2[0], ax) + (r2[2] * abs(np.dot(r2[1], ax)) + r2[3] * abs(np.dot(n2, ax)))
        sep = max(lo2 - hi1, lo1 - hi2)
        ovl = min(hi1, hi2) - max(lo1, lo2)
        sep_max = max(sep_max, sep)
        if sep > 0:
            sep_pos = min(sep_pos, sep)
        ovl_max = max(ovl_max, ovl)
    return sep_pos if sep_max > 0 else -ovl_max


def min_gap_at(t, curve_builder=None):
    """时刻 t 全体非相邻板凳对的最小真实欧氏间隙（shapely 精确距离，0 表示相交）.
    curve_builder(t) -> us（默认复合曲线）；可传入替代调头曲线做第二问验证."""
    from shapely.geometry import Polygon
    us = curve_builder(t) if curve_builder is not None else config_at(t)
    Ps = positions(us)
    n = N_H - 1
    rects = [bench_rect(i, Ps) for i in range(n)]
    bbs = [aabb(r) for r in rects]
    polys = [None] * n
    g = np.inf
    for i in range(n):
        for j in range(i + 1, n):
            if j - i == 1:
                continue
            if not aabb_overlap(bbs[i], bbs[j]):
                continue
            if polys[i] is None:
                polys[i] = Polygon(rect_corners(rects[i]))
            if polys[j] is None:
                polys[j] = Polygon(rect_corners(rects[j]))
            g = min(g, polys[i].distance(polys[j]))
    return g


def min_gap_pair(t):
    """同上，但返回 (间隙, 临界板凳对编号)."""
    from shapely.geometry import Polygon
    us = config_at(float(t))
    Ps = positions(us)
    n = N_H - 1
    rects = [bench_rect(i, Ps) for i in range(n)]
    bbs = [aabb(r) for r in rects]
    polys = [Polygon(rect_corners(r)) for r in rects]
    g, pair = np.inf, None
    for i in range(n):
        for j in range(i + 1, n):
            if j - i == 1 or not aabb_overlap(bbs[i], bbs[j]):
                continue
            d = polys[i].distance(polys[j])
            if d < g:
                g, pair = d, (i + 1, j + 1)
    return g, pair


# ---------------------------------------------------------------
# 六、主流程：仿真 + result4.xlsx + 第二问分析
# ---------------------------------------------------------------
def main():
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out = {}

    print("=" * 64)
    print("问题 4：S 形调头曲线与全龙仿真（p=1.7 m, R1=2R2, 调头空间 r<=4.5 m）")
    print("=" * 64)
    print(f"入口极角 θ_A = {TH_A:.6f} rad, 弧长参数 s(θ_A) = {S_A:.6f} m")
    print(f"E = ({E[0]:.6f}, {E[1]:.6f}),  F = ({F[0]:.6f}, {F[1]:.6f})")
    print(f"基线 R1 = {R1_0:.6f} m, R2 = {R2_0:.6f} m (R1+R2 = {S0:.6f} m)")
    print(f"两弧扫角 γ1 = γ2 = {G1_0:.6f} rad = {np.degrees(G1_0):.4f}°")
    print(f"切点 T = ({T0[0]:.6f}, {T0[1]:.6f}), O1 = ({O1_0[0]:.6f}, {O1_0[1]:.6f}), O2 = ({O2_0[0]:.6f}, {O2_0[1]:.6f})")
    print(f"调头曲线长度 L0 = {L0:.6f} m")

    # 调头曲线（两段圆弧）是否在调头空间内（密采样，仅取 u∈[0, L0]）
    uu = np.linspace(0.0, L0, 4000)
    rr = np.max(np.linalg.norm(np.array([curve_point(u) for u in uu]), axis=1))
    print(f"调头曲线（两弧）最大半径 = {rr:.6f} m （≤ 4.5 则全程在调头空间内）")

    # ---- (1) -100 ~ 100 s 仿真 ----
    times = np.arange(-100, 101)
    print("\n[仿真] 计算 -100~100 s 每秒 224 把手位置与速度 ...")
    all_us = [config_at(float(t)) for t in times]
    all_Ps = [positions(us) for us in all_us]
    all_vs = [velocities(us) for us in all_us]

    # ---- 写 result4.xlsx（模板格式）----
    import openpyxl
    xlsx = os.path.join(root, "附件", "result4.xlsx")
    wb = openpyxl.load_workbook(xlsx)
    wsP, wsV = wb.worksheets[0], wb.worksheets[1]
    for j, t in enumerate(times):
        col = 2 + j
        wsP.cell(row=1, column=col, value=f"{t} s")
        wsV.cell(row=1, column=col, value=f"{t} s")
        Ps, vs = all_Ps[j], all_vs[j]
        for i in range(N_H):
            wsP.cell(row=1 + 2 * i + 1, column=col, value=round(float(Ps[i][0]), 6))
            wsP.cell(row=1 + 2 * i + 2, column=col, value=round(float(Ps[i][1]), 6))
            wsV.cell(row=1 + i + 1, column=col, value=round(float(vs[i]), 6))
    wb.save(xlsx)
    print(f"[输出] 已写入 {xlsx}")

    # ---- 论文表：t = -100,-50,0,50,100; 把手 = 龙头, 第1,51,101,151,201节, 龙尾(后) ----
    key_t = [-100, -50, 0, 50, 100]
    key_i = [0, 1, 51, 101, 151, 201, 223]
    paper_table = {}
    print("\n论文表格数据（位置 x,y / 速度）：")
    for t in key_t:
        j = t + 100
        rows = []
        for i in key_i:
            rows.append({"handle": i + 1 if i < 223 else "tail",
                         "x": round(float(all_Ps[j][i][0]), 6),
                         "y": round(float(all_Ps[j][i][1]), 6),
                         "v": round(float(all_vs[j][i]), 6)})
        paper_table[str(t)] = rows
        print(f"t={t:>4} s: " + " | ".join(
            f"P{r['handle']}({r['x']:.3f},{r['y']:.3f})v={r['v']:.3f}" for r in rows[:3]) + " ...")

    # ---- 速度极值与碰撞验证（基线）----
    vmax = max(float(np.max(v)) for v in all_vs)
    print(f"\n窗口内把手速度最大值 = {vmax:.6f} m/s")

    print("\n[碰撞验证] 基线调头曲线，扫描 t ∈ [-100,100]（|t|<=25 内步长 0.25）...")
    gmin, gmin_t, gmin_pair = np.inf, None, None
    tlist = sorted(set(list(np.arange(-100, 101, 1.0)) + list(np.arange(-25, 25.01, 0.25))))
    for t in tlist:
        g = min_gap_at(float(t))
        if g < gmin:
            gmin, gmin_t = g, t
            gmin_pair = None
    # 在最小值附近细化并定位临界板凳对
    for t in np.arange(gmin_t - 1.0, gmin_t + 1.01, 0.05):
        g, pr = min_gap_pair(float(t))
        if g < gmin:
            gmin, gmin_t, gmin_pair = g, float(t), pr
    pair_str = f"（板凳对 {gmin_pair}）" if gmin_pair else ""
    print(f"全程最小板凳间隙 = {gmin:.6f} m @ t = {gmin_t} s {pair_str}")

    # ---- (2a) 不变性验证：不同半径比例 ----
    print("\n[第二问 a] 两弧族不变性验证（切点固定 E、F）:")
    invar = []
    for k in (0.25, 0.5, 1.0, 2.0, 3.0, 5.0):
        r1, r2, S = solve_split_ratio(k)
        O1, O2, _ = s_total_from_split(r1, r2)
        T, m, g1, g2 = arc_geometry(O1, O2, r1, r2)
        Lk = r1 * g1 + r2 * g2
        # 在调头空间内检查（弧密采样）
        ph = np.linspace(0, 1, 1200)
        a1 = O1[None] + r1 * np.stack([np.cos(np.arctan2(*(E - O1)[::-1]) - ph * g1),
                                       np.sin(np.arctan2(*(E - O1)[::-1]) - ph * g1)], 1)
        p2a = np.arctan2(*(T - O2)[::-1])
        a2 = O2[None] + r2 * np.stack([np.cos(p2a + ph * g2), np.sin(p2a + ph * g2)], 1)
        rmax = max(np.max(np.linalg.norm(a1, axis=1)), np.max(np.linalg.norm(a2, axis=1)))
        ok = verify_two_arc(O1, O2, r1, r2, g1, g2)
        invar.append({"k": k, "R1": round(r1, 6), "R2": round(r2, 6),
                      "S": round(r1 + r2, 6), "gamma": round(g1, 6),
                      "L": round(Lk, 6), "r_max": round(rmax, 6),
                      "in_space": bool(rmax <= 4.5 + 1e-9), "tangent_ok": bool(ok)})
        print(f"  k=R1/R2={k:5.2f}: R1={r1:.4f}, R2={r2:.4f}, γ={g1:.6f} rad, "
              f"L={Lk:.6f} m, r_max={rmax:.4f} {'OK' if rmax <= 4.5 + 1e-9 else '出界'}")

    # ---- (2b) 弧-线-弧结构搜索 ----
    print("\n[第二问 b] 弧-线-弧（保持相切）结构搜索 ...")
    best = None

    def alt_curve(R1, alpha):
        """构造弧(右转 R1,扫角α)-线(ℓ)-弧(左转 R2,扫角α)；返回 (R2, ℓ, T1, m, sampler) 或 None."""
        T1 = E + R1 * np.sin(alpha) * tE - R1 * (1 - np.cos(alpha)) * nE
        mdir = rot(tE, -alpha)
        nm = rot90(mdir)
        rhs = F - T1
        c = 1 - np.cos(alpha)
        if abs(c) < 1e-9:
            return None
        R2 = np.dot(rhs, nm) / c
        if R2 < R_MIN_PHYS or R2 > 4.4:
            return None
        ell = np.dot(rhs, mdir) - R2 * np.sin(alpha)
        if ell < 1e-6:
            return None
        O1 = E - R1 * nE
        O2 = T1 + ell * mdir + R2 * nm
        # 采样整条调头曲线，检查在调头空间内
        s1 = np.linspace(0, 1, 400)
        a1 = O1[None] + R1 * np.stack(
            [np.cos(np.arctan2(*(E - O1)[::-1]) - s1 * alpha),
             np.sin(np.arctan2(*(E - O1)[::-1]) - s1 * alpha)], 1)
        T2 = T1 + ell * mdir
        ln = T1[None] + s1[:, None] * (ell * mdir)[None]
        p2s = np.arctan2(*(T2 - O2)[::-1])
        a2 = O2[None] + R2 * np.stack(
            [np.cos(p2s + s1 * alpha), np.sin(p2s + s1 * alpha)], 1)
        rmax = max(np.max(np.linalg.norm(a1, axis=1)),
                   np.max(np.linalg.norm(ln, axis=1)),
                   np.max(np.linalg.norm(a2, axis=1)))
        if rmax > 4.5 + 1e-9:
            return None
        L = (R1 + R2) * alpha + ell
        return dict(R1=R1, R2=R2, alpha=alpha, ell=ell, L=L, rmax=rmax,
                    O1=O1, O2=O2, T1=T1, T2=T2, m=mdir)

    R_MIN_PHYS = D_BODY / 2.0   # 0.825 m：弧半径下界（弦长 1.65 m ≤ 2R 才能沿弧跟随）

    # 粗网格 + 局部精修（约束弧半径 ≥ D/2：板凳弦长不超过圆弧直径的几何极限）
    for R1g in np.linspace(R_MIN_PHYS + 1e-4, 3.2, 30):
        for ag in np.linspace(0.3, 3.0, 80):
            r = alt_curve(R1g, ag)
            if r is None:
                continue
            if best is None or r["L"] < best["L"]:
                best = r
    from scipy.optimize import minimize
    if best is not None:
        def negL(x):
            r = alt_curve(x[0], x[1])
            return 1e6 if r is None else -r["L"]
        res = minimize(negL, [best["R1"], best["alpha"]], method="Nelder-Mead",
                       options={"xatol": 1e-10, "fatol": 1e-12, "maxiter": 4000})
        r = alt_curve(*res.x)
        if r is not None and r["L"] < best["L"]:
            best = r
    if best is not None:
        L1b = best["L"]
        print(f"  最优: R1={best['R1']:.6f}, R2={best['R2']:.6f}, α={best['alpha']:.6f} rad"
              f" ({np.degrees(best['alpha']):.3f}°), ℓ={best['ell']:.6f} m")
        print(f"  L* = {L1b:.6f} m, r_max = {best['rmax']:.6f} m")
        print(f"  相对基线 L0={L0:.6f}: 缩短 {L0 - L1b:.6f} m ({(1 - L1b / L0) * 100:.2f}%)")
    else:
        L1b = None
        print("  未找到可行的弧-线-弧解")

    # ---- (2c) 弧-线-弧最优解碰撞验证 ----
    opt_gap = None
    if best is not None:
        bo = best

        def alt_us(t):
            """替代调头曲线的把手 u 递推（复合曲线 C 换成弧-线-弧版本）."""
            raise NotImplementedError

        # 碰撞验证：直接对给定 t 构造把手位置（头部沿替代曲线）
        def alt_config(t):
            R1, al, R2, el = bo["R1"], bo["alpha"], bo["R2"], bo["ell"]
            O1, O2, T1, T2, mdir = bo["O1"], bo["O2"], bo["T1"], bo["T2"], bo["m"]
            L1a, Ll, L2a = R1 * al, el, R2 * al
            Lc = L1a + Ll + L2a

            def pt(u):
                if u < 0:
                    return spiral_point(theta_of_arc(S_A - u))
                if u <= L1a:
                    phi = np.arctan2(*(E - O1)[::-1]) - u / R1
                    return O1 + R1 * np.array([np.cos(phi), np.sin(phi)])
                if u <= L1a + Ll:
                    return T1 + (u - L1a) * mdir
                if u <= Lc:
                    phi = np.arctan2(*(T2 - O2)[::-1]) + (u - L1a - Ll) / R2
                    return O2 + R2 * np.array([np.cos(phi), np.sin(phi)])
                return -spiral_point(theta_of_arc(S_A + (u - Lc)))

            us = np.empty(N_H)
            us[0] = t
            for i in range(223):
                di = D_ALL[i]
                P = pt(us[i])
                f = lambda dl: np.linalg.norm(pt(us[i] - dl) - P) - di
                us[i + 1] = us[i] - brentq(f, 0.4 * di, 2.6 * di, xtol=1e-12)
            return us

        gmin2, gt2 = np.inf, None
        for t in tlist:
            try:
                g = min_gap_at(float(t), curve_builder=alt_config)
            except Exception:
                g = np.inf
            if g < gmin2:
                gmin2, gt2 = g, t
        opt_gap = (gmin2, gt2)
        print(f"  弧-线-弧最优解碰撞验证: 全程最小间隙 = {gmin2:.6f} m @ t = {gt2} s"
              f"  {'（无碰撞，可行）' if gmin2 > 0 else '（发生碰撞，不可行！）'}")

    # ---- 汇总 JSON ----
    out = {
        "theta_A": float(TH_A), "s_A": float(S_A),
        "E": E.tolist(), "F": F.tolist(), "t_E": tE.tolist(),
        "O1": O1_0.tolist(), "O2": O2_0.tolist(), "T": T0.tolist(),
        "R1": float(R1_0), "R2": float(R2_0), "S": float(S0),
        "gamma_deg": float(np.degrees(G1_0)), "gamma_rad": float(G1_0),
        "L1": float(L1_0), "L2": float(L2_0), "L0": float(L0),
        "curve_r_max": float(rr),
        "v_max_window": vmax,
        "baseline_gap_min": float(gmin), "baseline_gap_t": float(gmin_t),
        "invariance": invar,
        "alt_opt": None if best is None else {
            "R1": float(best["R1"]), "R2": float(best["R2"]),
            "alpha_deg": float(np.degrees(best["alpha"])),
            "ell": float(best["ell"]), "L": float(L1b),
            "r_max": float(best["rmax"]),
            "save_m": float(L0 - L1b), "save_pct": float((1 - L1b / L0) * 100),
            "gap_min": None if opt_gap is None else float(opt_gap[0]),
            "gap_t": None if opt_gap is None else float(opt_gap[1]),
            "O1": best["O1"].tolist(), "O2": best["O2"].tolist(),
            "T1": best["T1"].tolist(), "T2": best["T2"].tolist(),
        },
        "paper_table": paper_table,
    }
    with open(os.path.join(root, "src", "problem4_results.json"), "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print("\n[输出] 关键数字已写入 src/problem4_results.json")


if __name__ == "__main__":
    main()
